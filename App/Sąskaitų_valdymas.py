from pathlib import Path
import tkinter as tk
from tkinter import Toplevel, filedialog, messagebox, StringVar
from tkinter import ttk
import customtkinter as ctk
import locale
import sys
from fpdf import FPDF
import datetime
import openpyxl
import smtplib
from email.message import EmailMessage
import os
import io
from num2words import num2words
import ctypes
import smtplib
from fuzzywuzzy import fuzz
from fuzzywuzzy import process


# Set DPI awareness for Windows (for better scaling)
ctypes.windll.shcore.SetProcessDpiAwareness(1)


# User class to store user data
class User:
    def __init__(self, nr=None, name=None, price=None, seria=None, pay_for=None, sum_in_words=None, email = None):
        self.nr = nr
        self.name = name
        self.price = price
        self.seria = seria
        self.pay_for = pay_for
        self.sum_in_words = sum_in_words
        self.email = email

main_user = ""
users = []
count = 0
lithuanian_months = [
        "Sausio", "Vasario", "Kovo", "Balandžio", "Gegužės", "Birželio",
        "Liepos", "Rugpjūčio", "Rugsėjo", "Spalio", "Lapkričio", "Gruodžio"
    ]
lithuanian_months_inaginink = [
    "Sausį", "Vasarį", "Kovą", "Balandį", "Gegužę", "Birželį",
        "Liepą", "Rugpjūtį", "Rugsėjį", "Spalį", "Lapkritį", "Gruodį"
]
today = datetime.date.today()
current_month_name = lithuanian_months[today.month - 1]
def number_to_words(number):
        if number is None:
            return "nulis eurų ir 00 centų"
        try:
            number = float(number)
        except ValueError:
            return "nulis eurų ir 00 centų"
        
        euros = int(number)
        cents = int(round((number - euros) * 100))

    # Convert euros and cents to words
        euros_words = num2words(euros, lang='lt')
        cents_words = num2words(cents, lang='lt')

        if euros == 100:
            euros_words = "šimtas"
        elif 100 < euros < 200:
            euros_words = euros_words.replace("vienas šimtas", "šimtas")

        last_digit_euros = euros % 10
        last_digit_cents = cents % 10

        if euros == 0 or euros % 10 == 0:
            euro_term = "eurų"
        elif last_digit_euros == 1:
            euro_term = "euras"
        elif 1 < last_digit_euros < 10:
            euro_term = "eurai"
        else:
            euro_term = "eurų"

        if cents == 0 or cents % 10 == 0:
            cent_term = "centų"
        elif last_digit_cents == 1:
            cent_term = "centas"
        elif 1 < last_digit_cents < 10:
            cent_term = "centai"
        else:
            cent_term = "centų"


    # Construct the final string:
            
        result = f"{euros_words} {euro_term}"
        if cents > 0:
            result += f" ir {cents_words} {cent_term}"
        else:
            result += " ir 00 centų"
        return result
def get_cell_value(cell):
    return cell.value
# Function to import users from the Excel file
def import_users(file_path):
    nr = 0
    global users
    users.clear()  # Clear the existing list of users before importing new ones
    
    missing_data_users = []  # List to hold users with missing data
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active  # Get the first sheet

        headers = {
            "Numeris": None,
            "Pirkėjas": None,
            "Kaina": None,
            "Serija": None,
            "Pavadinimas": None,
            "Gmail": None
        }

        actual_headers = [get_cell_value(cell) for cell in sheet[1]]

        # Fuzzy match headers
        for key in headers.keys():
            match = process.extractOne(key, actual_headers, scorer=fuzz.token_sort_ratio)
            if match and match[1] > 70:  # Use a threshold of 80 for similarity
                headers[key] = actual_headers.index(match[0])

        '''for cell in sheet[1]:
            cell_value = get_cell_value(cell)
            if cell_value in headers:
                headers[cell_value] = cell.column - 1  # store zero-based column index'''

        # Check if all headers are found
        missing_headers = [key for key, value in headers.items() if value is None]
        if missing_headers:
            messagebox.showwarning("Įspėjimas", f"Trūksta šių stulpelių: {', '.join(missing_headers)}\nGrįžtama prie pradinių duomenų pozicijų:Pirkėjas, Kaina, Serija, Pavadinimas, Gmail")

        # Default columns if headers are not found
        default_columns = {
            "Numeris": 0,
            "Pirkėjas": 1,
            "Kaina": 2,
            "Serija": 3,
            "Pavadinimas": 4,
            "Gmail": 5
        }

        # Use default columns if headers are missing
        for key, value in headers.items():
            if value is None:
                headers[key] = default_columns[key]

        # Start processing from the second row
        for row in sheet.iter_rows(min_row=2):
            # Skip completely blank rows
            if not any(get_cell_value(cell) for cell in row):
                continue

            # Extract data based on the found headers or default columns
            nr += 1
            name = get_cell_value(row[headers["Pirkėjas"]])
            price = get_cell_value(row[headers["Kaina"]])
            seria = get_cell_value(row[headers["Serija"]])
            pay_for = get_cell_value(row[headers["Pavadinimas"]])
            email = get_cell_value(row[headers["Gmail"]])
            sum_in_words = number_to_words(price)

            # Create a new user object
            user = User(nr, name, price, seria, pay_for, sum_in_words, email)

            # Check for missing data and add to missing_data_users list
            missing_data_fields = []
            if not nr:
                missing_data_fields.append('Numeris')
            if not name:
                missing_data_fields.append('Vardas')
            if not price:
                missing_data_fields.append('Kaina')
            if not seria:
                missing_data_fields.append('Serija')
            if not pay_for:
                missing_data_fields.append('Pavadinimas')
            if not email:
                missing_data_fields.append('gmail')

            if missing_data_fields:
                missing_data_users.append((user, missing_data_fields))
            else:
                users.append(user)  # Only add to users if all required fields are filled

        # After importing, inform the user about missing data
        if missing_data_users:
            missing_info = "\n".join(
                [f"Žmogui nr: {user.nr}, trūkta: {', '.join(fields)}" for user, fields in missing_data_users]
            )
            messagebox.showerror("Duomenų trūkumas", f"Kai kuriem žmonėm trūksta duomenų:\n{missing_info}")
        else:
            messagebox.showinfo("Importavimas sėkmingas", "Visi duomenys buvo rasti ir importuoti")

    except Exception as e:
        messagebox.showerror("Error", f"Error reading file: {e}")
    for user in users:
        safe_print(f"\nImported Data - Number:{user.nr}, Name: {user.name}, Price: {user.price}, Serie: {user.seria}, Pay For: {user.pay_for}, Email:{user.email}, Number in words: {user.sum_in_words}")    
# Function to get cell value from Excel
# PDF class for creating invoices
class PDF(FPDF):
    def header(self):
        self.set_font("DejaVu", 'B', 12)
        self.cell(0, 10, "SĄSKAITA FAKTŪRA", 0, 1, 'C')

    def footer(self):
        self.set_y(-15)
        self.set_font("DejaVu", 'I', 8)
        self.cell(0, 10, f'Page {self.page_no()}', 0, 0, 'C')

# Function to export user data to PDF
# PDF class for creating invoices
class PDF(FPDF):
    def header(self):
        self.set_font("Arial", 'B', 12)  # Changed to a built-in font
        self.cell(0, 10, "SĄSKAITA FAKTŪRA", 0, 1, 'C')

    def footer(self):
        self.set_y(-15)
        self.set_font("Arial", 'I', 8)  # Changed to a built-in font
        self.cell(0, 10, f'Page {self.page_no()}', 0, 0, 'C')

# PDF class for creating invoices
class PDF(FPDF):
    def header(self):
        self.set_font("DejaVu", 'B', 12)  # Use bold Unicode font
        self.cell(0, 10, "SĄSKAITA FAKTŪRA", 0, 1, 'C')

    def footer(self):
        self.set_y(-15)
        self.set_font("DejaVu", 'I', 8)  # Use italic Unicode font
        self.cell(0, 10, f'Page {self.page_no()}', 0, 0, 'C')

def export(pdf_file_path, user):

    app_directory = Path(__file__).parent.resolve()
    pdf_folder = app_directory / 'Sąskaitos'
    print(f"saskaitu path:{pdf_folder}")
    os.makedirs(pdf_folder, exist_ok=True)

    pdf = PDF()

    # Load the fonts
    pdf.add_font("DejaVu", "", "DejaVuSans.ttf", uni=True)  # Regular font
    pdf.add_font("DejaVu", "B", "DejaVuSans-Bold.ttf", uni=True)  # Bold font
    pdf.add_font("DejaVu", "I", "DejaVuSans-Oblique.ttf", uni=True)  # Italic font, if needed

    pdf.add_page()
    pdf.set_font("DejaVu", '', 12)  # Set to the regular Unicode font

    # Title and date
    pdf.cell(0, 10, f"Serija {user.seria[:9]} Nr. {user.seria[9:]}", 0, 1, 'C')
    pdf.cell(0, 10, f"{today.year} m. {lithuanian_months[today.month - 1]} {today.day} d.", 0, 1, 'C')

    pdf.ln(10)

    # Seller and Buyer Information
    pdf.set_font('DejaVu', 'B', 12)  # Bold font for header
    pdf.cell(90, 10, "Pardavėjas:", 0, 0)  # Seller header
    pdf.cell(0, 10, "Pirkėjas:", 0, 1)  # Buyer header

    pdf.set_font('DejaVu', '', 12)  # Regular font for buyer's name
    pdf.cell(90, 10, "", 0, 0)  # Empty cell to align buyer's name on the right
    pdf.cell(0, 10, user.name, 0, 1)  # Buyer name on the right

# Set font for seller's information
    pdf.set_font('DejaVu', '', 12)  # Regular font for content

# Seller Information
    pdf.cell(90, 10, "Capital Badminton Club VŠĮ", 0, 1)  # Seller name
    pdf.cell(90, 10, "Įm.k.: 301806862", 0, 1)  # Seller company number
    pdf.cell(90, 10, "Kaštonų g. 5 - 12 Vilnius", 0, 1)  # Seller address
    pdf.cell(90, 10, "A.s. LT722140030002649070", 0, 1)  # Seller account number
    pdf.cell(90, 10, "Luminor", 0, 1)  # Seller bank name
    pdf.cell(90, 10, "SWIFT: NDEALT2X", 0, 1)  # Seller SWIFT code
    pdf.cell(90, 10, "Registro tvarkytojas: Valstybės Įmonė Registrų Centras", 0, 1)  # Seller registry


# Add a line break after the tables

    pdf.ln(10)
    pdf.cell(0, 10, f"Apmokėti iki {today + datetime.timedelta(days=7)}", 0, 1)

    pdf.ln(10)

    # Invoice Table
    pdf.set_font('DejaVu', 'B', 8)  # Use bold Unicode font for table header
    col_width = pdf.w / 7.5
    th = pdf.font_size * 2

    pdf.cell(col_width * 0.5, th, "Eil. Nr.", border=1)
    pdf.cell(col_width * 0.5, th, "Kodas", border=1)
    pdf.cell(col_width * 3, th, "Pavadinimas", border=1)
    pdf.cell(col_width * 0.75, th, "Mato vnt.", border=1)
    pdf.cell(col_width * 0.75, th, "Kiekis", border=1)
    pdf.cell(col_width * 0.75, th, "Kaina Eur", border=1)
    pdf.cell(col_width * 0.75, th, "Suma Eur", border=1)
    pdf.ln(th)

    pdf.set_font('DejaVu', '', 10)  # Use regular Unicode font for table content
    pdf.cell(col_width * 0.5, th, "1", border=1)
    pdf.cell(col_width * 0.5, th, "111", border=1)

    #pdf.cell(col_width * 3, th, user.pay_for, border=1)
    x_before = pdf.get_x()
    y_before = pdf.get_y()

    pdf.multi_cell(col_width * 3, th, user.pay_for, border=1)
    x_after = pdf.get_x()
    y_after = pdf.get_y()
    pdf.set_xy(x_before + col_width * 3, y_before)
    cell_height = y_after - y_before

    #th = y_after - y_before

    pdf.cell(col_width * 0.75, cell_height, "Vnt.", border=1)
    pdf.cell(col_width * 0.75, cell_height, "1.00", border=1)
    pdf.cell(col_width * 0.75, cell_height, str(user.price), border=1)
    pdf.cell(col_width * 0.75, cell_height, str(user.price), border=1)
    pdf.ln(cell_height)

    pdf.ln(10)

    sum_in_words_col_width = len(user.sum_in_words)

    # Price Table
    pdf.set_font('DejaVu', 'B', 10)  # Header for Price Table
    pdf.cell(col_width * 2, th, "Bendra suma žodžiais:", border=1)
    pdf.cell(sum_in_words_col_width * 2.2, th, user.sum_in_words, border=1)
    pdf.ln(th)

    pdf.cell(col_width * 2, th, "Iš viso Eur:", border=1)
    pdf.cell(col_width * 3, th, str(float(user.price)) + "", border=1)
    pdf.ln(th)

    pdf.cell(col_width * 2, th, "Bendra suma Eur:", border=1)
    pdf.cell(col_width * 3, th, str(float(user.price)) + "", border=1)
    pdf.ln(th)

    pdf.ln(10)

    # Signatures
    pdf.set_font('DejaVu', '', 12)  # Use regular Unicode font for signatures
    pdf.cell(0, 10, "Sąskaitą išrašė, prekes išdavė:         ______ ______________________________________", 0, 1)
    pdf.set_font('DejaVu', '', 8)
    pdf.cell(0, 10, "                                                                                                       (pareigos, vardas, pavardė, parašas)", 0, 1)
    pdf.set_font('DejaVu', '', 12)
    pdf.cell(0, 10, "Sąskaitą gavo, prekes priėmė:           ______ ______________________________________", 0, 1)
    pdf.set_font('DejaVu', '', 8)
    pdf.cell(0, 10, "                                                                                                       (pareigos, vardas, pavardė, parašas)", 0, 1)

    # Save the PDF
    pdf.output(pdf_file_path)
    safe_print(f"PDF created successfully: {pdf_file_path}")

# Function to print Unicode safely
def safe_print(text):
    print(text.encode(sys.stdout.encoding, errors='replace').decode(sys.stdout.encoding))

# App window
class InvoiceApp:
    def __init__(self, master):
        self.master = master
        master.title("Sąskaitų generavimas")
        master.geometry("550x580")  # Adjust the main window size
        
        self.gmail = ""
        self.password = ""
        self.gmail_entry = None
        self.password_entry = None
        
        ctk.set_appearance_mode("Dark")
        ctk.set_default_color_theme("blue")
        
        self.file_path = None
        self.error_output = io.StringIO()
        
        # Create a main frame that fills the window
        self.main_frame = ctk.CTkFrame(master, width=600, height=600, corner_radius=10)
        self.main_frame.grid(row=0, column=0, padx=0, pady=0, sticky="nsew")
        
        # Prevent the frame from resizing based on content
        self.main_frame.columnconfigure(0, weight=1)
        self.main_frame.columnconfigure(1, weight=1)
        self.main_frame.rowconfigure(0, weight=1)
        self.main_frame.rowconfigure(1, weight=1)
        self.main_frame.rowconfigure(2, weight=1)
        self.main_frame.rowconfigure(3, weight=1)
        self.main_frame.rowconfigure(4, weight=1)
        self.main_frame.rowconfigure(5, weight=1)
        self.main_frame.rowconfigure(6, weight=1)

        self.user_label = ctk.CTkLabel(self.main_frame, text="", font=("Arial", 16))
        self.user_label.grid(row=0, column=0, padx=25, sticky="nw")

        self.night_mode_var = ctk.IntVar(value=1)
        self.dark_mode_checkbox = ctk.CTkCheckBox(self.main_frame, text="Naktinis rėžimas", variable=self.night_mode_var,command=self.toggle_dark_mode)
        self.dark_mode_checkbox.grid(row=0, column=0, padx=(150, 0), sticky="ne")

        # Create an import frame (shorter height)
        self.import_frame = ctk.CTkFrame(self.main_frame, width=300, height=150, corner_radius=10)
        self.import_frame.grid(row=0, column=0, padx=20, pady=(30, 0), sticky="nsew")
        
        # Create label for Excel file selection inside the import frame
        self.label = ctk.CTkLabel(self.import_frame, text="Pasirinkite Excel failą:", font=("Arial", 16))
        self.label.grid(row=0, column=0, padx=20, pady=(20, 10), sticky="w")

        # Create a textbox inside the import frame and place it inside the textbox frame
        self.imported_users_textbox = ctk.CTkTextbox(self.import_frame, height=150, width=200)
        self.imported_users_textbox.grid(row=1, column=0, padx=(200, 0), pady=(0, 30), sticky="n")

        # Import button inside the import frame
        self.import_button = ctk.CTkButton(self.import_frame, text="Importuoti", command=lambda: self.import_file(self.imported_users_textbox))
        self.import_button.grid(row=1, column=0, padx=20, pady=10, sticky="wn")

        # PDF export button
        self.export_button = ctk.CTkButton(self.main_frame, text="Generuoti PDF", command=self.export_pdf)
        self.export_button.grid(row=3, column=0, padx=20, pady=10, sticky="w")

        self.export_progress_bar = ttk.Progressbar(self.main_frame, orient="horizontal", length=200, mode="determinate")
        self.export_progress_bar.grid(row=3, column=0, padx=(300,10), pady=10, sticky="news")

        # Login section header
        self.login_label = ctk.CTkLabel(self.main_frame, text="Prisijunkite prie savo paskyros norėdami išsiųsti PDF failus:", font=("Arial", 14))
        self.login_label.grid(row=4, column=0, padx=20, pady=10, sticky='w')

        # Login button
        self.login_button = ctk.CTkButton(self.main_frame, text="Prisijungti", command=self.login)
        self.login_button.grid(row=5, column=0, padx=20, pady=10, sticky="w")

        # Write email button
        self.write_email_button = ctk.CTkButton(self.main_frame, text="Rašyti laišką", command=lambda: self.write_letter(self.user_label.cget("text")))
        self.write_email_button.grid(row=6, column=0, padx=20, pady=10, sticky="we")

    def login(self):

        login_window = Toplevel(self.master)
        login_window.title("Prisijungti")
        login_window.geometry("500x300")

        self.login_frame = ctk.CTkFrame(login_window)
        self.login_frame.pack(fill="both", expand="true")

        self.label = ctk.CTkLabel(self.login_frame, text="Prisijungimas", font=("Roboto", 24))
        self.label.pack(pady=12, padx=10)

        gmail_entry = ctk.CTkEntry(self.login_frame, placeholder_text="Gmail", width=200)
        gmail_entry.pack(pady=12, padx=10)

        password_entry = ctk.CTkEntry(self.login_frame, placeholder_text="Slaptažodis", show="*")
        password_entry.pack(pady=12, padx=10)

        prisijungti_button = ctk.CTkButton(self.login_frame, text="Prisijungti", command=lambda: self.take_gmail_password(gmail_entry, password_entry, login_window, self.user_label))
        prisijungti_button.pack(pady=12, padx=10)

    def write_letter(self, main_user):
        print(f"main_user: {main_user}")
        self.master.update_idletasks()
        write_email = tk.Toplevel(self.master)
        write_email.title("Laiškas")
        write_email.geometry("620x820")

        # Create a main frame to hold everything
        main_frame = ctk.CTkFrame(write_email)
        main_frame.grid(row=0, column=0, sticky="nsew")

        self.user_name = ctk.CTkLabel(main_frame, text=main_user, font=("Arial", 20))
        self.user_name.grid(row=0, column=0, padx=10, pady=(0, 20), sticky="nw")

        # Create the subject and body frame
        input_frame = ctk.CTkFrame(main_frame)
        input_frame.grid(row=0, column=0, padx=10, pady=(40, 0), sticky="nsew")

        # Configure rows and columns of the input frame
        input_frame.grid_rowconfigure(0, weight=1)
        input_frame.grid_rowconfigure(1, weight=2)  # Make the body section take more space
        input_frame.grid_columnconfigure(0, weight=1)
        input_frame.grid_columnconfigure(1, weight=2)

        # Subject Section (in its own frame inside the main frame)
        subject_frame = ctk.CTkFrame(input_frame)
        subject_frame.grid(row=1, column=0, columnspan=2, padx=10, pady=(20, 10), sticky="nsew")

        self.subject_label = ctk.CTkLabel(subject_frame, text="Tema:", font=("Arial", 20))
        self.subject_label.grid(row=0, column=0, padx=20, pady=(20, 0), sticky="e")
        
        self.subject_entry = ctk.CTkEntry(subject_frame, width=300)
        self.subject_entry.grid(row=0, column=1, padx=20, pady=(20, 5), sticky="w")

        # Body Section (in its own frame inside the main frame)
        body_frame = ctk.CTkFrame(input_frame)
        body_frame.grid(row=2, column=0, columnspan=2, padx=10, pady=20, sticky="nsew")

        self.body_label = ctk.CTkLabel(body_frame, text="Laiškas:", font=("Arial", 20))
        self.body_label.grid(row=0, column=0, padx=20, pady=5, sticky="ne")
        
        self.body_entry = ctk.CTkTextbox(body_frame, wrap="word", width=300, height=150)
        self.body_entry.grid(row=0, column=1, padx=20, pady=5, sticky="w")

        # Listbox and checkbox section
        listbox_frame = ctk.CTkFrame(main_frame)
        listbox_frame.grid(row=2, column=0, padx=10, pady=10, sticky="nsew")

        self.listbox_label = ctk.CTkLabel(listbox_frame, text="Pasirinkite kam siųsti:", font=("Arial", 16))
        self.listbox_label.grid(row=0, column=1, padx=20, pady=5, sticky="nw")

        user_selection_listbox = tk.Listbox(listbox_frame, selectmode=tk.MULTIPLE, width=25, height=10)
        user_selection_listbox.grid(row=2, column=1, padx=20, pady=5, sticky="e")
        self.update_user_listbox(user_selection_listbox)

        checkbox_people_var = tk.BooleanVar()
        self.checkbox_people = ctk.CTkCheckBox(listbox_frame, text="Pažymėti visus žmones", variable=checkbox_people_var, 
                                            command=lambda: self.toggle_selection_checkbox_people(user_selection_listbox, checkbox_people_var))
        self.checkbox_people.grid(row=2, column=0, padx=20, pady=(30, 0), sticky="w")

        self.progress = ttk.Progressbar(main_frame, orient="horizontal", length=200, mode="determinate")
        self.progress.grid(row=3, column=0, columnspan=2, pady=(10, 10))

        progress_bar_label = ctk.CTkLabel(main_frame, text=f"isšsiųsta:0/{len(users)}", font=("Arial", 12))
        progress_bar_label.grid(row=4, column=0, pady=(0, 20), sticky="new")

        # Gmail sending button
        self.gmail_button = ctk.CTkButton(main_frame, text="Siųsti gmail", command=lambda: self.send_gmail(user_selection_listbox, progress_bar_label))
        self.gmail_button.grid(row=4, column=0, columnspan=2, pady=(30, 5), sticky="s")

    def show_help(self):
        help_window = Toplevel(self.root)
        help_window.title("Pagalba")
        help_window.geometry("1200x300")

            
        # Add some text to the help window
        help_label = ctk.CTkLabel(help_window, text="Jei neveikia programėlė įsitikinkite:\n1. Stulpelių pavadinimai yra taisyklingi", justify="left", font=("Arial", 12))
        #help_label.configure(bg=bg_color, fg=fg_color)
        help_label.pack(pady=10, padx=20)

        # Load and display images
        try:
            correct_column_names = tk.PhotoImage(file="correct_column_names.png")  # Ensure the image path is correct
            help_window.correct_column_names = correct_column_names  # Keep a reference to avoid garbage collection
            correct_column_names_label = ctk.CTkLabel(help_window, image=correct_column_names)
            correct_column_names_label.pack(pady=10)
        except Exception as e:
            print(f"Error loading image: {e}")
            no_image_label = ctk.CTkLabel(help_window, text="Image not found", font=("Arial", 12))
            #no_image_label.configure(bg=bg_color, fg=fg_color)
            no_image_label.pack(pady=10)

        # Close button
        close_button = ctk.CTkButton(help_window, text="Close", command=help_window.destroy)
        close_button.pack(pady=5)

    def import_file(self, textbox):
        self.file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx;*.xls")])
        if self.file_path:
            textbox.delete("1.0", tk.END)
            import_users(self.file_path)
            #for index, user in enumerate(users):
                #self.imported_users_textbox.insert(tk.END, f"{index+1}: {user.name}\n")
            self.update_user_textbox(textbox)

    def update_user_textbox(self, box):
        box.delete("1.0", tk.END)
        for index, user in enumerate(users):
            box.insert(tk.END, f"{index+1}: {user.name}\n")
    
    def update_user_listbox(self, box):
        box.delete(0, tk.END)
        for index, user in enumerate(users):
            box.insert(tk.END, f"{index+1}: {user.name}\n")

    def export_pdf(self):
        
        custom_error_message = ''
        generated_pdf_list = "\n"
        failed_pdf_list = "\n"
        index = 0
        total_users = len(users)
        self.export_progress_bar["maximum"] = total_users 
        for user in users:
            try:
                if user.price is not None and user.seria is not None and user.pay_for is not None and user.email is not None:
                    pdf_file_path = f"{user.seria}{user.name}.pdf"
                    print(f"pdf file path in export func:{pdf_file_path}")
                    export(pdf_file_path, user)
                    index += 1
                    generated_pdf_list += f"{index}:{user.name}\n"
                    #add_to_dict(generated_pdf_list, user.name + "\n")
                else:
                    failed_pdf_list += user.name + "\n"
                    #add_to_dict(failed_pdf_list, user.name + "\n")
                    custom_error_message = "Klaida su kaikuriais duomenimis"
                
            except Exception as e:
                safe_print(f"Failed to create pdf for {user.name}: {e}")
                failed_pdf_list += user.name + "\n"
                generated_pdf_list = generated_pdf_list.replace(user.name + "\n", "")
                custom_error_message = "Klaida su duomenimis, peržiūrėkite juos"

            self.export_progress_bar["value"] = index
            self.master.update_idletasks()

        if custom_error_message:
            messagebox.showerror("Klaida", f"{custom_error_message}\nPDF failai nesukurti šiems žmonėms:{failed_pdf_list}\nSukurti šiems žmonėms:{generated_pdf_list}")
        else:
            messagebox.showinfo("Sėkmė!", f"PDF failai sukurti sėkmingai šiems žmonėms:{generated_pdf_list}")

        self.error_output.truncate(0)
        self.error_output.seek(0)
        self.export_progress_bar["value"] = 0

    def check_gmail_validity(self, gmail, password):
        try:
            # Attempt to connect to Gmail's SMTP server
            server = smtplib.SMTP_SSL("smtp.gmail.com", 465)
            server.login(gmail, password)
            server.quit()  # Quit the server after a successful login
            return 1
        except smtplib.SMTPAuthenticationError:
            return 2
        except Exception as e:
            error_message = str(e)
            if "getaddrinfo failed" in error_message:
                print(f"Error: {e}")
                return 3

    def take_gmail_password(self, gmailas, passwordas, login_window, label):
        
        #if self.gmail_entry and self.password_entry:# This function can be used to retrieve Gmail and password
            login_gmail = gmailas.get()
            login_password = passwordas.get()
            
        # Check if the email and password are valid
            if self.check_gmail_validity(login_gmail, login_password) == 2:
                messagebox.showerror("Klaida", "Neteisingi prisijungimo duomenys. Patikrinkite gmail arba slaptažodį.")
            elif self.check_gmail_validity(login_gmail, login_password) == 1:
                self.gmail = login_gmail
                self.password = login_password
                messagebox.showinfo("Sėkmė", "Prisijungimas sėkmingas!")
                main_user = f"Sveiki, {login_gmail}"
                label.configure(text=main_user)
                login_window.destroy()
            else:
                safe_print("No internet")
                messagebox.showerror("Klaida", "Nėra internetinio ryšio")

    def send_gmail(self, user_selection_listbox, label):
        selected_items = user_selection_listbox.curselection()  # Get selected users in the ListBox

        if selected_items:  # If some items are selected
            bad_signin_msg = ""
            failed_to_send_ppl_msg = ""
            sent_ppl_list = "\n"
            failed_ppl_list = "\n"
            selected_users_index = 0

            #login_gmail = self.gmail_entry.get()
            #login_password = self.password_entry.get()
            subject = self.subject_entry.get()
            body = self.body_entry.get("1.0", tk.END).strip()

            total_items = len(selected_items)
            self.progress["maximum"] = total_items 

            # Loop over the selected items (selected indices)
            for index in selected_items:
                selected_users_index += 1
                #print("selected users index:", selected_users_index)
                user = users[index]  # Get the selected user object from the users list

                if user.email:  # Check if the user has an email address
                    pdf_file_path = f"{user.seria}{user.name}.pdf"
                    try:
                        # Attempt to send the email with the PDF attachment
                        send_email(user.email, pdf_file_path, subject, body, self.gmail, self.password)
                        
                        sent_ppl_list += f"{user.nr}: {user.email}\n"
                        safe_print(f"PDF file:{pdf_file_path} sent to {user.email}")
                    except Exception as e:
                        error_message = str(e)
                        safe_print(f"Failed to send:{pdf_file_path} to {user.email}: {e}")
                        
                        if "Username and Password not accepted" in error_message or "BadCredentials" in error_message:
                            bad_signin_msg = "Neteisingi prisijungimo duomenys. Patikrinkite gmail arba slaptažodį."
                            break
                        elif f"The recipient address <{user.email}> is not a valid RFC 5321 address":
                            
                            failed_ppl_list += f"{user.nr}: {user.email}\n"
                            # Generic error message for other failures
                            failed_to_send_ppl_msg = "Failai nebuvo nusiųsti šiais adresais:"

                self.progress["value"] = selected_users_index  # Update the progress bar
                #self.master.update_idletasks()
                label.configure(text=f"Išsiųsta:{selected_users_index}/{total_items}")
                self.master.update_idletasks()

            # Show the appropriate message box based on success or failure
            if bad_signin_msg:
                messagebox.showerror("Klaida", bad_signin_msg)
            elif failed_to_send_ppl_msg:
                messagebox.showinfo("Nusiųsta", f"{failed_to_send_ppl_msg}\n {failed_ppl_list}\nTačiau buvo nusiųsti šiais:\n{sent_ppl_list}")

            else:
                messagebox.showinfo("Sėkmė!", f"PDF failai nusiųsti sėkmingai šiais adresais:{sent_ppl_list}")

            # Reset error output buffer
            self.error_output.truncate(0)  
            self.error_output.seek(0)
            self.progress["value"] = 0
        else:
            # If no users were selected, show an error message
            messagebox.showerror("Klaida", "Prašome pasirinkti kam siųsti")

    def toggle_selection_checkbox_people(self, user_selection_listbox, checkbox_people_var):
        if checkbox_people_var.get():
        # Select all items in the Listbox
            user_selection_listbox.select_set(0, tk.END)
        else:
        # Deselect all items in the Listbox
            user_selection_listbox.select_clear(0, tk.END)
    def toggle_dark_mode(self):
        if self.night_mode_var.get() == 1:
            ctk.set_appearance_mode("Dark") 
            print("Dark mode enabled")
        else:
            ctk.set_appearance_mode("Light")
            print("Dark mode disabled")
def send_email(to_email, pdf_path, subject, body, login_gmail, login_password):
    msg = EmailMessage()
    msg['Subject'] = subject
    msg['From'] = login_gmail
    msg['To'] = to_email
    msg.set_content(body)

    with open(pdf_path, 'rb') as f:
        file_data = f.read()
        file_name = os.path.basename(pdf_path)
    
    msg.add_attachment(file_data, maintype='application', subtype='pdf', filename=file_name)
    
    # Send the email
    with smtplib.SMTP_SSL('smtp.gmail.com', 465) as smtp:
        smtp.login(login_gmail, login_password)
        smtp.send_message(msg)

def show_main_app():
    #startup_window.destroy()
    root = tk.Tk()
    app = InvoiceApp(root)
    root.mainloop()
    
'''    def show_help(self, dark_mode):
        help_window = Toplevel(self.root)
        help_window.title("Pagalba")
        help_window.geometry("1200x300")

        if dark_mode:
            ctk.set_appearance_mode("Dark")
        else:
            ctk.set_appearance_mode("Light")
            
        # Add some text to the help window
        help_label = ctk.CTkLabel(help_window, text="Jei neveikia programėlė įsitikinkite:\n1. Stulpelių pavadinimai yra taisyklingi", justify="left", font=("Arial", 12))
        #help_label.configure(bg=bg_color, fg=fg_color)
        help_label.pack(pady=10, padx=20)

        # Load and display images
        try:
            correct_column_names = tk.PhotoImage(file="correct_column_names.png")  # Ensure the image path is correct
            help_window.correct_column_names = correct_column_names  # Keep a reference to avoid garbage collection
            correct_column_names_label = ctk.CTkLabel(help_window, image=correct_column_names)
            correct_column_names_label.pack(pady=10)
        except Exception as e:
            print(f"Error loading image: {e}")
            no_image_label = ctk.CTkLabel(help_window, text="Image not found", font=("Arial", 12))
            #no_image_label.configure(bg=bg_color, fg=fg_color)
            no_image_label.pack(pady=10)

        # Close button
        close_button = ctk.CTkButton(help_window, text="Close", command=help_window.destroy)
        close_button.pack(pady=5)'''

    
            
# Main function to run the app
def main():
    root = tk.Tk()
    app = InvoiceApp(root)
    root.mainloop()

if __name__ == "__main__":
    main()


